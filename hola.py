import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
from datetime import datetime

# Clase Usuario
class Usuario:
    def __init__(self, user_id, nombre, membresia):
        self.user_id = user_id
        self.nombre = nombre
        self.membresia = membresia
        self.pagos = 0
        self.entradas = 0
        self.invitados = []

# Clase Membresia
class Membresia:
    def __init__(self, tipo):
        self.tipo = tipo
        self.beneficios = self.definir_beneficios()

    def definir_beneficios(self):
        if self.tipo == "corriente":
            return ["Acceso a la instalación", "Acceso a los dispositivos", "Acceso a un entrenador"]
        elif self.tipo == "plus":
            return ["Acceso a la instalación", "Acceso a los dispositivos", "Acceso a un entrenador", "Ingreso limitado a invitados (7 ingresos al mes)"]
        else:
            return []

# Clase Invitado
class Invitado:
    def __init__(self, nombre, user_id):
        self.nombre = nombre
        self.user_id = user_id

# Clase GymSystem
class GymSystem:
    def __init__(self, db_path='gym_data.xlsx'):
        self.db_path = db_path
        self.init_db()

    def init_db(self):
        if not os.path.exists(self.db_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Usuarios'
            sheet.append(['ID', 'Nombre', 'Membresía', 'Pagos', 'Entradas', 'Invitados'])
            workbook.save(self.db_path)

    def register_user(self, user_id, nombre, membresia):
        try:
            workbook = openpyxl.load_workbook(self.db_path)
            sheet = workbook['Usuarios']
            sheet.append([user_id, nombre, membresia, 0, 0, ""])
            workbook.save(self.db_path)
        except Exception as e:
            raise e

    def get_user(self, user_id):
        try:
            workbook = openpyxl.load_workbook(self.db_path)
            sheet = workbook['Usuarios']
            for row in sheet.iter_rows(values_only=True):
                if row[0] == user_id:
                    return Usuario(row[0], row[1], row[2])
            return None
        except Exception as e:
            raise e

    def update_user(self, user):
        try:
            workbook = openpyxl.load_workbook(self.db_path)
            sheet = workbook['Usuarios']
            for row in sheet.iter_rows():
                # Interfaz de registro
                self.label_id = tk.Label(root, text="ID del Usuario")
                self.label_id.pack()
                self.entry_id = tk.Entry(root)
                self.entry_id.pack()

                self.label_name = tk.Label(root, text="Nombre del Usuario")
                self.label_name.pack()
                self.entry_name = tk.Entry(root)
                self.entry_name.pack()

                self.label_membership = tk.Label(root, text="Tipo de Membresía (corriente/plus)")
                self.label_membership.pack()
                self.entry_membership = tk.Entry(root)
                self.entry_membership.pack()

                self.btn_register = tk.Button(root, text="Registrar Usuario", command=self.register_user)
                self.btn_register.pack()

                self.btn_activate_membership = tk.Button(root, text="Activar Membresía",
                                                         command=self.activate_membership)
                self.btn_activate_membership.pack()

                self.btn_register_guest = tk.Button(root, text="Registrar Invitado", command=self.register_guest)
                self.btn_register_guest.pack()

            def register_user(self):
                user_id = self.entry_id.get()
                nombre = self.entry_name.get()
                membresia = self.entry_membership.get()

                try:
                    self.system.register_user(user_id, nombre, membresia)
                    messagebox.showinfo("Éxito", "Usuario registrado correctamente")
                except Exception as e:
                    messagebox.showerror("Error", f"Error al registrar usuario: {e}")

            def activate_membership(self):
                user_id = self.entry_id.get()
                user = self.system.get_user(user_id)
                if user:
                    user.membresia = "activa"
                    self.system.update_user(user)
                    messagebox.showinfo("Éxito", "Membresía activada correctamente")
                else:
                    messagebox.showerror("Error", "Usuario no encontrado")

            def register_guest(self):
                user_id = self.entry_id.get()
                guest_name = self.entry_name.get()
                user = self.system.get_user(user_id)
                if user:
                    if user.membresia == "plus" and len(user.invitados) < 7:
                        invitado = Invitado(guest_name, user_id)
                        user.invitados.append(invitado)
                        self.system.update_user(user)
                        messagebox.showinfo("Éxito", "Invitado registrado correctamente")
                    else:
                        messagebox.showerror("Error", "No se puede registrar más invitados o membresía no es plus")
                else:
                    messagebox.showerror("Error", "Usuario no encontrado")

            if __name__ == "__main__":
                root = tk.Tk()
                app = GymApp(root)
                root.mainloop()
                if row[0].value == user.user_id:
                    row[3].value = user.pagos
                    row[4].value = user.entradas
                    row[5].value = ",".join([inv.nombre for inv in user.invitados])
                    break
            workbook.save(self.db_path)
        except Exception as e:
            raise e

# Clase GymApp
class GymApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema del Gimnasio")
        self.system = GymSystem()
